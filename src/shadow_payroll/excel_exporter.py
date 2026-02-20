"""
Excel export functionality for shadow payroll reports.

This module handles generating formatted Excel reports with
proper styling and currency formatting. Supports both single-scenario
legacy reports and multi-scenario comparison workbooks.
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Any, Dict

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pathlib import Path

from .models import ShadowPayrollResult
from .scenarios import normalize_line_items

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Handles Excel report generation for shadow payroll results.

    Creates professionally formatted Excel files with proper
    currency formatting and cell alignment.
    """

    # Styling constants
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    CURRENCY_FORMAT = '[$ARS] #,##0.00'
    COLUMN_WIDTH_LABEL = 40
    COLUMN_WIDTH_VALUE = 80

    @staticmethod
    def create_report(result: ShadowPayrollResult) -> BytesIO:
        """
        Generate Excel report from shadow payroll results.

        Args:
            result: Complete shadow payroll calculation result

        Returns:
            BytesIO: Excel file as bytes buffer ready for download

        Example:
            >>> exporter = ExcelExporter()
            >>> excel_bytes = exporter.create_report(result)
            >>> # Can be used with Streamlit download button
        """
        logger.info("Generating Excel report")

        # Convert result to display dictionary
        data = result.to_display_dict()

        # Create DataFrame
        df = pd.DataFrame(list(data.items()), columns=["Field", "Value"])

        # Create Excel file in memory
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Shadow Payroll')

            workbook = writer.book
            worksheet = writer.sheets['Shadow Payroll']

            # Apply styling
            ExcelExporter._apply_styling(worksheet, df)

        output.seek(0)
        logger.info("Excel report generated successfully")

        return output

    @staticmethod
    def _apply_styling(worksheet, df: pd.DataFrame) -> None:
        """
        Apply formatting and styling to worksheet.

        Args:
            worksheet: openpyxl worksheet object
            df: DataFrame containing the data
        """
        # Set column widths
        worksheet.column_dimensions['A'].width = ExcelExporter.COLUMN_WIDTH_LABEL
        worksheet.column_dimensions['B'].width = ExcelExporter.COLUMN_WIDTH_VALUE

        # Style header row
        for cell in worksheet[1]:
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Apply text wrapping and alignment to all cells
        align = Alignment(wrap_text=True, vertical='top', horizontal='left')
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = align

        # Apply ARS currency format to numeric values (rows 2-7)
        # These correspond to the monetary fields in the result
        monetary_fields = [
            "Gross Monthly (ARS)",
            "Income Tax Monthly (est.)",
            "Employee Contributions",
            "Net Employee",
            "Employer Contributions",
            "Total Employer Cost",
        ]

        for row_idx in range(2, len(df) + 2):  # +2 because Excel is 1-indexed + header
            cell_label = worksheet[f"A{row_idx}"].value
            if cell_label in monetary_fields:
                worksheet[f"B{row_idx}"].number_format = ExcelExporter.CURRENCY_FORMAT

        logger.debug("Styling applied to Excel worksheet")

    @staticmethod
    def create_detailed_report(
        result: ShadowPayrollResult, input_summary: Dict[str, Any]
    ) -> BytesIO:
        """
        Generate detailed Excel report with multiple sheets.

        Args:
            result: Shadow payroll calculation result
            input_summary: Dictionary with input parameters and additional context

        Returns:
            BytesIO: Excel file with multiple sheets

        Note:
            Creates two sheets:
            1. Summary - Main results
            2. Details - Input parameters and assumptions
        """
        logger.info("Generating detailed Excel report")

        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = result.to_display_dict()
            df_summary = pd.DataFrame(
                list(summary_data.items()), columns=["Field", "Value"]
            )
            df_summary.to_excel(writer, index=False, sheet_name='Summary')

            # Details sheet
            df_details = pd.DataFrame(
                list(input_summary.items()), columns=["Parameter", "Value"]
            )
            df_details.to_excel(writer, index=False, sheet_name='Details')

            # Apply styling to both sheets
            workbook = writer.book
            ExcelExporter._apply_styling(workbook['Summary'], df_summary)
            ExcelExporter._apply_styling(workbook['Details'], df_details)

        output.seek(0)
        logger.info("Detailed Excel report generated successfully")

        return output


    # Logo path (resolved from package directory)
    LOGO_PATH = Path(__file__).parent / "logo.png"

    @classmethod
    def _add_logo(cls, ws: Any, cell: str = "A1") -> None:
        """Add company logo to a worksheet if the logo file exists."""
        if cls.LOGO_PATH.is_file():
            try:
                from openpyxl.drawing.image import Image as XlImage

                img = XlImage(str(cls.LOGO_PATH))
                img.width = 100
                img.height = 100
                ws.add_image(img, cell)
            except Exception:
                pass  # Skip logo if image insertion fails

    # --- Multi-scenario comparison styles ---
    BLUE_FILL = PatternFill(start_color="2d8cf0", end_color="2d8cf0", fill_type="solid")
    WHITE_BOLD_FONT = Font(bold=True, color="FFFFFF", size=11)
    COURIER_FONT = Font(name="Courier New", size=10)
    COURIER_BOLD = Font(name="Courier New", size=10, bold=True)
    GREEN_FILL = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    RED_FILL = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    TITLE_FONT = Font(bold=True, color="2d8cf0", size=14)
    THIN_BORDER_TOP = Border(top=Side(style="thin"))

    @staticmethod
    def _ensure_dict_line_items(scenarios: list[dict]) -> list[dict]:
        """Convert list-format line_items to dict format for normalize_line_items.

        Also populates line_items_full with local currency amounts when
        converting from list format.
        """
        prepared = []
        for s in scenarios:
            result = s.get("result", {})
            li = result.get("line_items", {})
            if isinstance(li, list):
                s = dict(s)
                s["result"] = dict(result)
                s["result"]["line_items"] = {
                    item["label"]: item.get("amount_usd", 0.0)
                    for item in li
                    if isinstance(item, dict)
                }
                s["result"]["line_items_full"] = {
                    item["label"]: {
                        "amount_usd": item.get("amount_usd", 0.0),
                        "amount_local": item.get("amount_local", 0.0),
                        "local_currency": item.get("local_currency", ""),
                    }
                    for item in li
                    if isinstance(item, dict)
                }
            prepared.append(s)
        return prepared

    def create_comparison_report(
        self,
        scenarios: list[dict],
        metadata: dict[str, Any] | None = None,
    ) -> BytesIO:
        """Generate multi-sheet Excel workbook with comparison and per-scenario sheets.

        Args:
            scenarios: List of ScenarioData dicts from session state.
            metadata: Optional dict with model_name and timestamp.

        Returns:
            BytesIO: Excel file ready for download.
        """
        metadata = metadata or {}
        scenarios = self._ensure_dict_line_items(scenarios)
        output = BytesIO()

        from openpyxl import Workbook

        wb = Workbook()

        # --- Comparison sheet ---
        ws_comp = wb.active
        ws_comp.title = "Comparison"
        self._build_comparison_sheet(ws_comp, scenarios, metadata)

        # --- Per-scenario sheets ---
        for scenario in scenarios:
            name = scenario.get("name", "Scenario")[:31]  # Excel 31-char limit
            ws = wb.create_sheet(title=name)
            self._build_scenario_sheet(ws, scenario, metadata)

        wb.save(output)
        output.seek(0)
        logger.info("Comparison Excel report generated (%d scenarios)", len(scenarios))
        return output

    def _build_comparison_sheet(
        self, ws: Any, scenarios: list[dict], metadata: dict[str, Any]
    ) -> None:
        """Build the Comparison sheet with styled comparison table."""
        num_scenarios = len(scenarios)

        # Logo in top-right corner
        last_col_letter = get_column_letter(1 + num_scenarios)
        self._add_logo(ws, f"{last_col_letter}1")

        # Title row (merged)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + num_scenarios)
        title_cell = ws.cell(row=1, column=1, value="Slater Consulting \u2014 Shadow Payroll Comparison")
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center")

        # Subtitle row
        model = metadata.get("model_name", "")
        ts = metadata.get("timestamp", datetime.now().strftime("%Y-%m-%d"))
        subtitle = f"Generated: {ts}"
        if model:
            subtitle += f" | Model: {model}"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1 + num_scenarios)
        sub_cell = ws.cell(row=2, column=1, value=subtitle)
        sub_cell.font = Font(size=9, color="888888")
        sub_cell.alignment = Alignment(horizontal="center")

        # Blank row
        start_row = 4

        # Header row
        labels, matrix = normalize_line_items(scenarios)
        header = ["Cost Category"] + [s.get("name", "?") for s in scenarios]

        for col_idx, val in enumerate(header, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=val)
            cell.fill = self.BLUE_FILL
            cell.font = self.WHITE_BOLD_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")

        # Data rows
        for label_idx, label in enumerate(labels):
            row_num = start_row + 1 + label_idx
            ws.cell(row=row_num, column=1, value=label)

            row_values = [matrix[s_idx][label_idx] for s_idx in range(num_scenarios)]
            min_val = min(row_values)
            max_val = max(row_values)
            all_equal = min_val == max_val

            for s_idx, val in enumerate(row_values):
                cell = ws.cell(row=row_num, column=2 + s_idx, value=val)
                cell.number_format = '#,##0'
                cell.font = self.COURIER_FONT
                cell.alignment = Alignment(horizontal="right")
                if not all_equal:
                    if val == min_val:
                        cell.fill = self.GREEN_FILL
                    elif val == max_val:
                        cell.fill = self.RED_FILL

        # Total row
        total_row = start_row + 1 + len(labels)
        ws.cell(row=total_row, column=1, value="Total Employer Cost").font = Font(bold=True)

        totals = [
            s.get("result", {}).get("total_employer_cost_usd", 0) for s in scenarios
        ]
        min_total = min(totals)
        max_total = max(totals)
        totals_equal = min_total == max_total

        for s_idx, val in enumerate(totals):
            cell = ws.cell(row=total_row, column=2 + s_idx, value=val)
            cell.number_format = '#,##0'
            cell.font = self.COURIER_BOLD
            cell.alignment = Alignment(horizontal="right")
            cell.border = self.THIN_BORDER_TOP
            if not totals_equal:
                if val == min_total:
                    cell.fill = self.GREEN_FILL
                elif val == max_total:
                    cell.fill = self.RED_FILL

        # Also add border to label cell
        ws.cell(row=total_row, column=1).border = self.THIN_BORDER_TOP

        # Auto-fit column widths
        ws.column_dimensions["A"].width = 25
        for col_idx in range(2, 2 + num_scenarios):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    def _build_scenario_sheet(
        self, ws: Any, scenario: dict, metadata: dict[str, Any]
    ) -> None:
        """Build a per-scenario detail sheet."""
        name = scenario.get("name", "Scenario")
        inp = scenario.get("input_data", {})
        result = scenario.get("result", {})

        # Logo + Title
        self._add_logo(ws, "C1")
        ws.cell(row=1, column=1, value=name).font = Font(bold=True, size=14, color="2d8cf0")

        # Input summary section
        row = 3
        ws.cell(row=row, column=1, value="Assignment Details").font = Font(bold=True, size=11)
        row += 1

        input_fields = [
            ("Annual Salary (USD)", f"${inp.get('salary_usd', 0):,.0f}"),
            ("Duration (months)", inp.get("duration_months", "N/A")),
            ("Home Country", inp.get("home_country", "N/A")),
            ("Host Country", inp.get("host_country", "N/A")),
            ("Housing Benefit (USD)", f"${inp.get('housing_usd', 0):,.0f}"),
            ("Education Benefit (USD)", f"${inp.get('school_usd', 0):,.0f}"),
            ("Spouse", "Yes" if inp.get("has_spouse") else "No"),
            ("Children", inp.get("num_children", 0)),
        ]

        for label, value in input_fields:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        # Cost breakdown
        ws.cell(row=row, column=1, value="Cost Breakdown").font = Font(bold=True, size=11)
        row += 1

        local_currency = result.get("local_currency", "Local")
        cost_headers = ["Cost Item", "USD", local_currency]
        for col_idx, val in enumerate(cost_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = self.BLUE_FILL
            cell.font = self.WHITE_BOLD_FONT
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
        row += 1

        line_items = result.get("line_items", {})
        line_items_full = result.get("line_items_full", {})
        if isinstance(line_items, dict):
            for label, amount_usd in line_items.items():
                ws.cell(row=row, column=1, value=label)
                cell_usd = ws.cell(row=row, column=2, value=float(amount_usd))
                cell_usd.number_format = '#,##0'
                cell_usd.font = self.COURIER_FONT
                cell_usd.alignment = Alignment(horizontal="right")
                # Use full details if available for local currency amount
                full = line_items_full.get(label, {})
                local_val = full.get("amount_local", 0) if full else 0
                cell_local = ws.cell(row=row, column=3, value=float(local_val) if local_val else 0)
                cell_local.number_format = '#,##0'
                cell_local.font = self.COURIER_FONT
                cell_local.alignment = Alignment(horizontal="right")
                row += 1

        # Total row
        total_usd = result.get("total_employer_cost_usd", 0)
        total_local = result.get("total_employer_cost_local", 0)
        ws.cell(row=row, column=1, value="Total Employer Cost").font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=total_usd)
        cell.number_format = '#,##0'
        cell.font = self.COURIER_BOLD
        cell.border = self.THIN_BORDER_TOP
        cell = ws.cell(row=row, column=3, value=total_local)
        cell.number_format = '#,##0'
        cell.font = self.COURIER_BOLD
        cell.border = self.THIN_BORDER_TOP
        ws.cell(row=row, column=1).border = self.THIN_BORDER_TOP
        row += 2

        # Cost rating
        cost_rating = result.get("cost_rating") or result.get("overall_rating")
        if isinstance(cost_rating, dict):
            rating_level = cost_rating.get("overall_rating", cost_rating.get("level", "N/A"))
            ws.cell(row=row, column=1, value="Cost Rating").font = Font(bold=True)
            ws.cell(row=row, column=2, value=rating_level)
            row += 2

        # AI Insights
        insights = result.get("insights_text") or result.get("insights_paragraph", "")
        if insights:
            ws.cell(row=row, column=1, value="AI Insights").font = Font(bold=True)
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            cell = ws.cell(row=row, column=1, value=str(insights))
            cell.alignment = Alignment(wrap_text=True)

        # Auto-fit columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18


def export_to_excel(result: ShadowPayrollResult) -> BytesIO:
    """
    Convenience function to export result to Excel.

    Args:
        result: Shadow payroll calculation result

    Returns:
        BytesIO: Excel file ready for download
    """
    exporter = ExcelExporter()
    return exporter.create_report(result)
