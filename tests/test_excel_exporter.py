"""
Unit tests for Excel export module.

Tests report generation, column headers, currency formatting,
and multi-sheet detailed reports.
"""

from io import BytesIO

import openpyxl
import pytest

from shadow_payroll.excel_exporter import ExcelExporter, export_to_excel
from shadow_payroll.models import (
    BaseCalculation,
    TaxCalculation,
    ShadowPayrollResult,
)


@pytest.fixture
def sample_result(sample_base_calculation, sample_tax_calculation):
    """Fixture providing a complete ShadowPayrollResult."""
    return ShadowPayrollResult(
        base=sample_base_calculation,
        tax=sample_tax_calculation,
        fx_date="2025-01-20",
        fx_source="open.er-api.com",
    )


class TestCreateReport:
    """Test suite for ExcelExporter.create_report / export_to_excel."""

    def test_returns_valid_excel_bytesio(self, sample_result):
        """Test create_report returns a valid BytesIO with xlsx content."""
        excel_bytes = export_to_excel(sample_result)
        assert isinstance(excel_bytes, BytesIO)

        # Verify it's a valid xlsx by loading it
        wb = openpyxl.load_workbook(excel_bytes)
        assert wb.active is not None
        wb.close()

    def test_sheet_name_is_shadow_payroll(self, sample_result):
        """Test the sheet is named 'Shadow Payroll'."""
        excel_bytes = export_to_excel(sample_result)
        wb = openpyxl.load_workbook(excel_bytes)
        assert wb.active.title == "Shadow Payroll"
        wb.close()

    def test_english_column_headers(self, sample_result):
        """Test that column headers are 'Field' and 'Value'."""
        excel_bytes = export_to_excel(sample_result)
        wb = openpyxl.load_workbook(excel_bytes)
        ws = wb.active

        assert ws["A1"].value == "Field"
        assert ws["B1"].value == "Value"
        wb.close()

    def test_monetary_fields_have_currency_format(self, sample_result):
        """Test that monetary value cells have ARS currency format."""
        excel_bytes = export_to_excel(sample_result)
        wb = openpyxl.load_workbook(excel_bytes)
        ws = wb.active

        # Find a monetary field row (e.g., "Gross Monthly (ARS)")
        currency_format_found = False
        for row_idx in range(2, ws.max_row + 1):
            label = ws[f"A{row_idx}"].value
            if label == "Gross Monthly (ARS)":
                fmt = ws[f"B{row_idx}"].number_format
                assert "ARS" in fmt or "#,##0" in fmt
                currency_format_found = True
                break

        assert currency_format_found, "No monetary field with currency format found"
        wb.close()

    def test_all_display_dict_fields_present(self, sample_result):
        """Test all fields from to_display_dict appear in the report."""
        expected_fields = list(sample_result.to_display_dict().keys())
        excel_bytes = export_to_excel(sample_result)
        wb = openpyxl.load_workbook(excel_bytes)
        ws = wb.active

        actual_fields = []
        for row_idx in range(2, ws.max_row + 1):
            val = ws[f"A{row_idx}"].value
            if val:
                actual_fields.append(val)

        for field in expected_fields:
            assert field in actual_fields, f"Missing field: {field}"
        wb.close()


class TestCreateDetailedReport:
    """Test suite for ExcelExporter.create_detailed_report."""

    def test_produces_two_sheets(self, sample_result):
        """Test detailed report has Summary and Details sheets."""
        input_summary = {
            "Salary USD": 100000.0,
            "Duration": "12 months",
            "FX Rate": 1000.0,
        }
        excel_bytes = ExcelExporter.create_detailed_report(sample_result, input_summary)
        wb = openpyxl.load_workbook(excel_bytes)

        assert "Summary" in wb.sheetnames
        assert "Details" in wb.sheetnames
        assert len(wb.sheetnames) == 2
        wb.close()

    def test_details_sheet_contains_input_params(self, sample_result):
        """Test Details sheet contains the input parameters."""
        input_summary = {
            "Salary USD": 100000.0,
            "Duration": "12 months",
        }
        excel_bytes = ExcelExporter.create_detailed_report(sample_result, input_summary)
        wb = openpyxl.load_workbook(excel_bytes)
        ws = wb["Details"]

        # Check header
        assert ws["A1"].value == "Parameter"
        assert ws["B1"].value == "Value"

        # Check data
        params = [ws[f"A{r}"].value for r in range(2, ws.max_row + 1)]
        assert "Salary USD" in params
        assert "Duration" in params
        wb.close()
